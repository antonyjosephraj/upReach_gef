import openpyxl

file = 'D:\\upReach\\Automation\\20230613 Testcases.xlsx'
workbook = openpyxl.load_workbook(file)

sheet = workbook["GEF"]

rows = sheet.max_row
cols = sheet.max_column

expected_output = ''
flag = 'PASS'
input_values = {}

for r in range(15, 20):
    for c in range(1, 10):
        if c == 6:
            aa = sheet.cell(r, c).value
            if aa != None:
                expected_output = sheet.cell(r, c+1).value
                lines = aa.split('\n')

                output = {}
                for line in lines:
                    key, value = line.split('=')

                    key = key.strip()
                    value = value.strip()

                    input_values[key] = value

print("input_values['First name']", input_values['First name'], input_values['Last name'])
