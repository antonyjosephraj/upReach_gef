import sys
sys.path.append("C:\\Users\\AntonyJosephRaj\\OneDrive - JMAN Group Ltd\\PROJECTS\\UPREACH\\Automation\\20230613_automation_v3")
from selenium import webdriver
from pageObjects.firstPage import FirstPage
from selenium.webdriver.common.by import By
import logging
from utilities.loggerFile import LogGen
import unittest
import HtmlTestRunner
import openpyxl
import time
import pytest

class Test_001_firstPage(unittest.TestCase):
    baseURL = "https://uat.getemployable.org/gef-assessment"

    # firstName = 'Antony'
    # lastName = 'Joseph Raj'
    # universityEmail = "antony@ac.uk"
    # personalEmail = "anotny@gmail.com"
    # mobileNumber = "07667876556"
    # dateOfBirth = "11-02-2023"
    file = 'D:\\upReach\\Automation\\20230613 Testcases.xlsx'
    workbook = openpyxl.load_workbook(file)

    sheet = workbook["GEF"]

    rows = sheet.max_row
    cols = sheet.max_column

    expected_output = ''
    flag = 'PASS'
    input_values = {}

    for r in range(15, 20):
        for c in range(1, 10):
            if c == 6:
                aa = sheet.cell(r, c).value
                if aa != None:
                    expected_output = sheet.cell(r, c+1).value
                    lines = aa.split('\n')

                    output = {}
                    for line in lines:
                        key, value = line.split('=')

                        key = key.strip()
                        value = value.strip()

                        input_values[key] = value

    print("input_values['First name']", input_values['First name'], input_values['Last name'])

    driver = webdriver.Chrome()
    logger = LogGen.loggen()


    @classmethod
    def setUpClass(cls):
        cls.driver.get(cls.baseURL)
        cls.driver.maximize_window()


    def test_enterFirstPage(self):
        self.driver.get(self.baseURL)
        self.logger.info("=========== Started First page title test =========")

        self.test_data = FirstPage(self.driver)

        self.logger.info("=========== Enter First name and Last name input values =========")

        if 'First name' in self.input_values:
            self.test_data.setFirstName(self.input_values['First name'])
        if 'Last name' in self.input_values:
            self.test_data.setLastName(self.input_values['Last name'])
        if 'University Email' in self.input_values:
            self.test_data.setUniversityEmail(self.input_values['University Email'])
        if 'Personal Email' in self.input_values:
            self.test_data.setPersonalEmail(self.input_values['Personal Email'])
        if 'Mobile number' in self.input_values:
            self.test_data.setMobileNumber(self.input_values['Mobile number'])
        if 'Date of birth' in self.input_values:
            self.test_data.setDateOfBirth(self.input_values['Date of birth'])
        if 'Gender' in self.input_values:
            self.test_data.setGender(self.input_values['Gender'])
        time.sleep(5)

        self.test_data.clickGotoSecondPage()
        time.sleep(5)

        error_message = ''
        your_university = self.driver.find_element(By.XPATH, "/html/body/div[1]/div/div[2]/div/div/div/form/fieldset[2]/h2").text
        assert self.expected_output.lower() == your_university.lower(), "Throughing Error..........."

        if (self.expected_output.lower() != your_university.lower()):
            error_message = self.driver.find_element(By.NAME, "errors").text

        time.sleep(5)

        if error_message != '':
            if "valid first name" in error_message:
                self.driver.save_screenshot(".\\screenshots\\" + "InvalidFirstNameFieldError.png")
            if "valid last name" in error_message:
                self.driver.save_screenshot(".\\screenshots\\" + "InvalidLastNameFieldError.png")
            if "valid university email" in error_message:
                self.driver.save_screenshot(".\\screenshots\\" + "InvalidUniversityEmailFieldError.png")
            if "valid personal email" in error_message:
                self.driver.save_screenshot(".\\screenshots\\" + "InvalidPersonalEmailFieldError.png")
            if "valid 10 digit mobile number" in error_message:
                self.driver.save_screenshot(".\\screenshots\\" + "InvalidMobileNumberFieldError.png")
            if "enter your date of birth" in error_message:
                self.driver.save_screenshot(".\\screenshots\\" + "InvalidDateOfBirthFieldError.png")
            if "choose your gender" in error_message:
                self.driver.save_screenshot(".\\screenshots\\" + "InvalidGenderFieldError.png")

        time.sleep(5)

        for r in range(15, 20):
            for c in range(1, 10):
                if c == 6:
                    aa = self.sheet.cell(r, c).value
                    if aa != None:
                        self.logger.info("=========== Error 1111111 =========")
                        self.sheet.cell(r, 9).value = self.flag
        self.workbook.save(self.file)

    @classmethod
    def tearDownClass(cls):
        cls.driver.close()

# python testCases/aboutYou.py
# pytest -v -s testCases/aboutYou.py


if __name__ == '__main__':
    unittest.main(testRunner=HtmlTestRunner.HTMLTestRunner(output='C:\\Users\\AntonyJosephRaj\\OneDrive - JMAN Group Ltd\\PROJECTS\\UPREACH\\Automation\\20230613_automation_v3\\reports'))
    # pytest.main(["-v", "--html=report.html"])